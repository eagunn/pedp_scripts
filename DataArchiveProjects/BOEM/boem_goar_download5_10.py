"""
BOEM Gulf of America Region (GOAR) Data Download Script
Downloads all available GIS and spatial data from BOEM's Gulf region.
Creates a comprehensive data catalog in Excel format.
"""

import os
import requests
from datetime import datetime
import json
from pathlib import Path
from urllib.parse import urljoin
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class BOEMDataDownloader:
    def __init__(self, output_dir):
        """
        Initialize the downloader with output directory.
        
        Args:
            output_dir (str): Path to external hard drive or destination folder
        """
        self.output_dir = Path(output_dir)
        self.base_url = "https://gis.boem.gov/arcgis/rest/services"
        self.data_portal_url = "https://www.data.boem.gov"
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        
        # Catalog to track all downloaded datasets
        self.data_catalog = []
        
        # Create main directory structure
        self.create_directory_structure()
        
    def create_directory_structure(self):
        """Create organized folder structure for downloads."""
        folders = [
            'cadastral',
            'leases',
            'infrastructure',
            'boundaries',
            'planning_areas',
            'protraction_diagrams',
            'rest_services',
            'metadata'
        ]
        
        for folder in folders:
            (self.output_dir / folder).mkdir(parents=True, exist_ok=True)
        
        print(f"Created directory structure at: {self.output_dir}")
    
    def add_to_catalog(self, dataset_name, category, file_path, format_type, description, source_url="", metadata=None):
        """Add a dataset entry to the catalog."""
        file_size = ""
        if os.path.exists(file_path):
            size_bytes = os.path.getsize(file_path)
            if size_bytes < 1024:
                file_size = f"{size_bytes} B"
            elif size_bytes < 1024 * 1024:
                file_size = f"{size_bytes / 1024:.2f} KB"
            elif size_bytes < 1024 * 1024 * 1024:
                file_size = f"{size_bytes / (1024 * 1024):.2f} MB"
            else:
                file_size = f"{size_bytes / (1024 * 1024 * 1024):.2f} GB"
        
        catalog_entry = {
            'Dataset Name': dataset_name,
            'Category': category,
            'Description': description,
            'File Path': str(file_path.relative_to(self.output_dir)),
            'Format': format_type,
            'File Size': file_size,
            'Download Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'Source URL': source_url,
            'Coordinate System': metadata.get('spatialReference', {}).get('wkid', 'Unknown') if metadata else 'Unknown',
            'Geometry Type': metadata.get('geometryType', 'Unknown') if metadata else 'Unknown',
            'Feature Count': metadata.get('count', 'Unknown') if metadata else 'Unknown'
        }
        
        self.data_catalog.append(catalog_entry)
    
    def download_file(self, url, output_path, description="file"):
        """Download a file with progress indication."""
        try:
            print(f"Downloading {description}...")
            response = self.session.get(url, stream=True, timeout=300)
            response.raise_for_status()
            
            total_size = int(response.headers.get('content-length', 0))
            
            with open(output_path, 'wb') as f:
                if total_size == 0:
                    f.write(response.content)
                else:
                    downloaded = 0
                    for chunk in response.iter_content(chunk_size=8192):
                        if chunk:
                            f.write(chunk)
                            downloaded += len(chunk)
                            progress = (downloaded / total_size) * 100
                            print(f"  Progress: {progress:.1f}%", end='\r')
            
            print(f"\n  ✓ Saved to: {output_path}")
            return True
            
        except Exception as e:
            print(f"\n  ✗ Error downloading {description}: {str(e)}")
            return False
    
    def get_arcgis_service_info(self, service_url):
        """Get information about an ArcGIS REST service."""
        try:
            params = {'f': 'json'}
            response = self.session.get(service_url, params=params, timeout=30)
            response.raise_for_status()
            return response.json()
        except Exception as e:
            print(f"Error getting service info: {str(e)}")
            return None
    
    def download_layer_data(self, service_url, layer_id, layer_name, output_folder):
        """Download data from a specific ArcGIS layer in preferred formats."""
        try:
            layer_url = f"{service_url}/{layer_id}"
            query_url = f"{layer_url}/query"
            
            base_params = {
                'where': '1=1',
                'outFields': '*',
                'returnGeometry': 'true'
            }
            
            print(f"  Querying layer: {layer_name}")
            safe_name = layer_name.replace(' ', '_').replace('/', '_')
            format_downloaded = False
            
            # Priority 1: Try File Geodatabase (preferred)
            try:
                params = base_params.copy()
                params['f'] = 'filegdb'
                print(f"    Attempting File Geodatabase format...")
                response = self.session.get(query_url, params=params, timeout=300)
                if response.status_code == 200 and len(response.content) > 0:
                    gdb_file = output_folder / f"{safe_name}.gdb.zip"
                    with open(gdb_file, 'wb') as f:
                        f.write(response.content)
                    print(f"    ✓ Saved File Geodatabase to: {gdb_file}")
                    format_downloaded = True
                    
                    # Add to catalog
                    self.add_to_catalog(
                        dataset_name=layer_name,
                        category='REST Service Layer',
                        file_path=gdb_file,
                        format_type='File Geodatabase',
                        description=f"Feature layer from {service_url.split('/')[-2]} service. Contains spatial and attribute data for {layer_name}.",
                        source_url=layer_url
                    )
            except Exception as e:
                print(f"    File Geodatabase not available: {str(e)}")
            
            # Priority 2: Try Shapefile
            if not format_downloaded:
                try:
                    params = base_params.copy()
                    params['f'] = 'shapefile'
                    print(f"    Attempting Shapefile format...")
                    response = self.session.get(query_url, params=params, timeout=300)
                    if response.status_code == 200 and len(response.content) > 0:
                        shp_file = output_folder / f"{safe_name}_shp.zip"
                        with open(shp_file, 'wb') as f:
                            f.write(response.content)
                        print(f"    ✓ Saved Shapefile to: {shp_file}")
                        format_downloaded = True
                        
                        # Add to catalog
                        self.add_to_catalog(
                            dataset_name=layer_name,
                            category='REST Service Layer',
                            file_path=shp_file,
                            format_type='Shapefile',
                            description=f"Feature layer from {service_url.split('/')[-2]} service. Contains spatial and attribute data for {layer_name}.",
                            source_url=layer_url
                        )
                except Exception as e:
                    print(f"    Shapefile not available: {str(e)}")
            
            # Priority 3: Try KML/KMZ
            if not format_downloaded:
                try:
                    params = base_params.copy()
                    params['f'] = 'kmz'
                    print(f"    Attempting KMZ format...")
                    response = self.session.get(query_url, params=params, timeout=300)
                    if response.status_code == 200 and len(response.content) > 0:
                        kmz_file = output_folder / f"{safe_name}.kmz"
                        with open(kmz_file, 'wb') as f:
                            f.write(response.content)
                        print(f"    ✓ Saved KMZ to: {kmz_file}")
                        format_downloaded = True
                        
                        # Add to catalog
                        self.add_to_catalog(
                            dataset_name=layer_name,
                            category='REST Service Layer',
                            file_path=kmz_file,
                            format_type='KMZ',
                            description=f"Feature layer from {service_url.split('/')[-2]} service. Contains spatial data for {layer_name} in Google Earth format.",
                            source_url=layer_url
                        )
                except Exception as e:
                    print(f"    KMZ not available: {str(e)}")
            
            # Priority 4: Fall back to GeoJSON (always available)
            if not format_downloaded:
                try:
                    params = base_params.copy()
                    params['f'] = 'geojson'
                    print(f"    Attempting GeoJSON format...")
                    response = self.session.get(query_url, params=params, timeout=300)
                    response.raise_for_status()
                    geojson_file = output_folder / f"{safe_name}.geojson"
                    with open(geojson_file, 'w') as f:
                        json.dump(response.json(), f, indent=2)
                    print(f"    ✓ Saved GeoJSON to: {geojson_file}")
                    format_downloaded = True
                    
                    # Add to catalog
                    self.add_to_catalog(
                        dataset_name=layer_name,
                        category='REST Service Layer',
                        file_path=geojson_file,
                        format_type='GeoJSON',
                        description=f"Feature layer from {service_url.split('/')[-2]} service. Contains spatial and attribute data for {layer_name} in web-friendly format.",
                        source_url=layer_url
                    )
                except Exception as e:
                    print(f"    GeoJSON error: {str(e)}")
            
            # Priority 5: Last resort - JSON
            if not format_downloaded:
                try:
                    params = base_params.copy()
                    params['f'] = 'json'
                    print(f"    Attempting JSON format...")
                    response = self.session.get(query_url, params=params, timeout=300)
                    response.raise_for_status()
                    json_file = output_folder / f"{safe_name}.json"
                    with open(json_file, 'w') as f:
                        json.dump(response.json(), f, indent=2)
                    print(f"    ✓ Saved JSON to: {json_file}")
                    format_downloaded = True
                    
                    # Add to catalog
                    self.add_to_catalog(
                        dataset_name=layer_name,
                        category='REST Service Layer',
                        file_path=json_file,
                        format_type='JSON',
                        description=f"Feature layer from {service_url.split('/')[-2]} service. Contains attribute data for {layer_name}.",
                        source_url=layer_url
                    )
                except Exception as e:
                    print(f"    JSON error: {str(e)}")
            
            return format_downloaded
            
        except Exception as e:
            print(f"  ✗ Error downloading layer {layer_name}: {str(e)}")
            return False
    
    def download_rest_services(self):
        """Download data from ArcGIS REST services."""
        print("\n=== Downloading from ArcGIS REST Services ===\n")
        
        services = [
            {
                'name': 'BOEM_BSEE MMC Layers',
                'url': 'https://gis.boem.gov/arcgis/rest/services/BOEM_BSEE/MMC_Layers/MapServer'
            },
            {
                'name': 'Gulf of Mexico',
                'url': 'https://gis.boem.gov/arcgis/rest/services/BOEM_BSEE/GOM_Layers/MapServer'
            }
        ]
        
        for service in services:
            print(f"\nProcessing service: {service['name']}")
            service_info = self.get_arcgis_service_info(service['url'])
            
            if not service_info:
                continue
            
            # Create folder for this service
            service_folder = self.output_dir / 'rest_services' / service['name'].replace(' ', '_')
            service_folder.mkdir(exist_ok=True)
            
            # Save service metadata
            with open(service_folder / 'service_info.json', 'w') as f:
                json.dump(service_info, f, indent=2)
            
            # Download each layer
            if 'layers' in service_info:
                for layer in service_info['layers']:
                    layer_id = layer['id']
                    layer_name = layer['name']
                    
                    self.download_layer_data(
                        service['url'],
                        layer_id,
                        layer_name,
                        service_folder
                    )
                    
                    time.sleep(1)  # Be respectful to the server
    
    def download_cadastral_data(self):
        """Download cadastral data in preferred formats (geodatabase, shapefile, etc)."""
        print("\n=== Downloading Cadastral Data ===\n")
        
        # BOEM data files with format preferences and descriptions
        cadastral_files = [
            {
                'name': 'Active_Leases',
                'description': 'Polygons representing active oil and gas leases in the Gulf of America. Includes lease numbers, operators, effective dates, and lease status. Critical for identifying current leasing activity and operator information.',
                'urls': [
                    'https://www.data.boem.gov/Mapping/Files/ActiveLeasePolygons.gdb.zip',
                    'https://www.data.boem.gov/Mapping/Files/ActiveLeasePolygons.zip',
                    'https://www.data.boem.gov/Mapping/Files/ActiveLeasePolygons_shp.zip'
                ],
                'folder': 'leases',
                'category': 'Leasing'
            },
            {
                'name': 'Block_Polygons',
                'description': 'OCS block boundaries representing the cadastral grid system used for leasing. Each block is approximately 5,760 acres (3 miles x 3 miles). Contains block numbers, protraction areas, and official boundary coordinates.',
                'urls': [
                    'https://www.data.boem.gov/Mapping/Files/BlockPolygons.gdb.zip',
                    'https://www.data.boem.gov/Mapping/Files/BlockPolygons.zip',
                    'https://www.data.boem.gov/Mapping/Files/BlockPolygons_shp.zip'
                ],
                'folder': 'cadastral',
                'category': 'Cadastral'
            },
            {
                'name': 'Protraction_Diagrams',
                'description': 'Protraction diagram boundaries that group OCS blocks into management units. Each protraction is typically 1 degree latitude by 2 degrees longitude. Used for planning and administrative purposes.',
                'urls': [
                    'https://www.data.boem.gov/Mapping/Files/ProtractionPolygons.gdb.zip',
                    'https://www.data.boem.gov/Mapping/Files/ProtractionPolygons.zip',
                    'https://www.data.boem.gov/Mapping/Files/ProtractionPolygons_shp.zip'
                ],
                'folder': 'protraction_diagrams',
                'category': 'Cadastral'
            },
            {
                'name': 'Oil_Gas_Platforms',
                'description': 'Point locations of oil and gas platforms/structures in the Gulf. Includes structure names, complex IDs, installation dates, water depths, and operational status. Essential for infrastructure mapping and spatial planning.',
                'urls': [
                    'https://www.data.boem.gov/Mapping/Files/Platforms.gdb.zip',
                    'https://www.data.boem.gov/Mapping/Files/Platforms.zip',
                    'https://www.data.boem.gov/Mapping/Files/Platforms_shp.zip',
                    'https://www.data.boem.gov/Mapping/Files/Platforms.kml'
                ],
                'folder': 'infrastructure',
                'category': 'Infrastructure'
            },
            {
                'name': 'Pipelines',
                'description': 'Pipeline routes and right-of-ways in the Gulf. Includes pipeline segments, operators, diameters, product types, and status. Critical for understanding offshore transportation networks and potential conflicts.',
                'urls': [
                    'https://www.data.boem.gov/Mapping/Files/Pipelines.gdb.zip',
                    'https://www.data.boem.gov/Mapping/Files/Pipelines.zip',
                    'https://www.data.boem.gov/Mapping/Files/Pipelines_shp.zip',
                    'https://www.data.boem.gov/Mapping/Files/Pipelines.kml'
                ],
                'folder': 'infrastructure',
                'category': 'Infrastructure'
            },
            {
                'name': 'Lease_Term_Lines',
                'description': 'Lines delineating 5-year and 10-year lease term durations based on water depth. Leases seaward of the line have 10-year terms (deeper water), while leases landward have 5-year terms (shallow water). Critical for understanding lease duration and planning exploration/development timelines. May vary by lease sale.',
                'urls': [
                    'https://www.data.boem.gov/Mapping/Files/LeaseTermLines.gdb.zip',
                    'https://www.data.boem.gov/Mapping/Files/LeaseTermLines.zip',
                    'https://www.data.boem.gov/Mapping/Files/LeaseTermLines_shp.zip',
                    'https://www.data.boem.gov/Mapping/Files/Lease_Term_Lines.zip'
                ],
                'folder': 'cadastral',
                'category': 'Cadastral'
            }
        ]
        
        for file_info in cadastral_files:
            downloaded = False
            for url in file_info['urls']:
                file_ext = url.split('.')[-1] if '.' in url.split('/')[-1] else 'zip'
                if url.endswith('.gdb.zip'):
                    file_ext = 'gdb.zip'
                
                output_path = self.output_dir / file_info['folder'] / f"{file_info['name']}.{file_ext}"
                
                if self.download_file(url, output_path, f"{file_info['name']} ({file_ext})"):
                    # Add to catalog
                    format_name = 'File Geodatabase' if file_ext == 'gdb.zip' else \
                                 'Shapefile' if 'shp' in file_ext or file_ext == 'zip' else \
                                 'KML' if file_ext == 'kml' else file_ext.upper()
                    
                    self.add_to_catalog(
                        dataset_name=file_info['name'],
                        category=file_info['category'],
                        file_path=output_path,
                        format_type=format_name,
                        description=file_info['description'],
                        source_url=url
                    )
                    downloaded = True
                    break
                else:
                    print(f"  Trying alternate format...")
            
            if not downloaded:
                print(f"  ⚠ Could not download {file_info['name']} in any format")
            
            time.sleep(2)
    
    def download_boundary_data(self):
        """Download boundary data in preferred formats."""
        print("\n=== Downloading Boundary Data ===\n")
        
        boundary_files = [
            {
                'name': 'State_Seaward_Boundaries',
                'description': 'Submerged Lands Act (SLA) boundaries marking the division between state and federal waters. Typically 3 nautical miles from shore (9 miles for Texas and Gulf Coast Florida). Critical for jurisdictional determinations.',
                'urls': [
                    'https://www.data.boem.gov/Mapping/Files/Boundaries.gdb.zip',
                    'https://www.data.boem.gov/Mapping/Files/Boundaries.zip',
                    'https://www.data.boem.gov/Mapping/Files/Boundaries_shp.zip',
                    'https://www.data.boem.gov/Mapping/Files/Boundaries.kml'
                ],
                'folder': 'boundaries',
                'category': 'Administrative Boundaries'
            },
            {
                'name': 'Planning_Areas',
                'description': 'BOEM planning area boundaries used for lease sale planning and resource management. Includes Western, Central, and Eastern Gulf planning areas. Used for regional analysis and program planning.',
                'urls': [
                    'https://www.data.boem.gov/Mapping/Files/PlanningAreas.gdb.zip',
                    'https://www.data.boem.gov/Mapping/Files/PlanningAreas.zip',
                    'https://www.data.boem.gov/Mapping/Files/PlanningAreas_shp.zip',
                    'https://www.data.boem.gov/Mapping/Files/PlanningAreas.kml'
                ],
                'folder': 'planning_areas',
                'category': 'Administrative Boundaries'
            },
            {
                'name': 'Fairways',
                'description': 'Designated navigation fairways and shipping channels in the Gulf. Areas where oil and gas activities may be restricted or prohibited to ensure safe vessel passage. Important for maritime planning and safety.',
                'urls': [
                    'https://www.data.boem.gov/Mapping/Files/Fairways.gdb.zip',
                    'https://www.data.boem.gov/Mapping/Files/Fairways.zip',
                    'https://www.data.boem.gov/Mapping/Files/Fairways_shp.zip'
                ],
                'folder': 'boundaries',
                'category': 'Maritime'
            },
            {
                'name': '8g_Zone',
                'description': 'The 8(g) revenue sharing zone extending 3 miles seaward of state waters (3-6 miles offshore). States receive 27% of revenues from leases in this zone. Important for fiscal and policy analysis.',
                'urls': [
                    'https://www.data.boem.gov/Mapping/Files/8g_Zone.gdb.zip',
                    'https://www.data.boem.gov/Mapping/Files/8g_Zone.zip',
                    'https://www.data.boem.gov/Mapping/Files/8g_Zone_shp.zip'
                ],
                'folder': 'boundaries',
                'category': 'Administrative Boundaries'
            }
        ]
        
        for file_info in boundary_files:
            downloaded = False
            for url in file_info['urls']:
                file_ext = url.split('.')[-1] if '.' in url.split('/')[-1] else 'zip'
                if url.endswith('.gdb.zip'):
                    file_ext = 'gdb.zip'
                
                output_path = self.output_dir / file_info['folder'] / f"{file_info['name']}.{file_ext}"
                
                if self.download_file(url, output_path, f"{file_info['name']} ({file_ext})"):
                    format_name = 'File Geodatabase' if file_ext == 'gdb.zip' else \
                                 'Shapefile' if 'shp' in file_ext or file_ext == 'zip' else \
                                 'KML' if file_ext == 'kml' else file_ext.upper()
                    
                    self.add_to_catalog(
                        dataset_name=file_info['name'],
                        category=file_info['category'],
                        file_path=output_path,
                        format_type=format_name,
                        description=file_info['description'],
                        source_url=url
                    )
                    downloaded = True
                    break
                else:
                    print(f"  Trying alternate format...")
            
            if not downloaded:
                print(f"  ⚠ Could not download {file_info['name']} in any format")
            
            time.sleep(2)
    
    def create_excel_catalog(self):
        """Create a comprehensive Excel data catalog."""
        print("\n=== Creating Excel Data Catalog ===\n")
        
        catalog_path = self.output_dir / 'BOEM_GOAR_Data_Catalog.xlsx'
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data Catalog"
        
        # Define headers
        headers = [
            'Dataset Name',
            'Category',
            'Description',
            'File Path',
            'Format',
            'File Size',
            'Download Date',
            'Source URL',
            'Coordinate System',
            'Geometry Type',
            'Feature Count'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Add data
        for row_num, entry in enumerate(self.data_catalog, 2):
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = entry.get(header, '')
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Adjust column widths
        column_widths = {
            'A': 30,  # Dataset Name
            'B': 20,  # Category
            'C': 60,  # Description
            'D': 40,  # File Path
            'E': 18,  # Format
            'F': 12,  # File Size
            'G': 20,  # Download Date
            'H': 50,  # Source URL
            'I': 18,  # Coordinate System
            'J': 15,  # Geometry Type
            'K': 15   # Feature Count
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Add auto-filter
        ws.auto_filter.ref = ws.dimensions
        
        # Create summary sheet
        summary_ws = wb.create_sheet("Summary")
        summary_ws.append(['BOEM Gulf of America Region Data Catalog'])
        summary_ws.append([''])
        summary_ws.append(['Download Information'])
        summary_ws.append(['Download Date:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        summary_ws.append(['Total Datasets:', len(self.data_catalog)])
        summary_ws.append(['Output Directory:', str(self.output_dir)])
        summary_ws.append([''])
        summary_ws.append(['Datasets by Category'])
        
        # Count by category
        category_counts = {}
        format_counts = {}
        for entry in self.data_catalog:
            cat = entry.get('Category', 'Unknown')
            fmt = entry.get('Format', 'Unknown')
            category_counts[cat] = category_counts.get(cat, 0) + 1
            format_counts[fmt] = format_counts.get(fmt, 0) + 1
        
        for category, count in sorted(category_counts.items()):
            summary_ws.append([category, count])
        
        summary_ws.append([''])
        summary_ws.append(['Datasets by Format'])
        for fmt, count in sorted(format_counts.items()):
            summary_ws.append([fmt, count])
        
        # Style summary sheet
        summary_ws['A1'].font = Font(bold=True, size=14)
        summary_ws['A3'].font = Font(bold=True, size=12)
        summary_ws['A8'].font = Font(bold=True, size=12)
        summary_ws.column_dimensions['A'].width = 30
        summary_ws.column_dimensions['B'].width = 15
        
        # Create metadata sheet with field descriptions
        metadata_ws = wb.create_sheet("Field Descriptions")
        metadata_ws.append(['Field Name', 'Description'])
        
        field_descriptions = [
            ['Dataset Name', 'Name of the spatial dataset or layer'],
            ['Category', 'Thematic category (e.g., Cadastral, Infrastructure, Boundaries)'],
            ['Description', 'Detailed description of dataset contents, use cases, and key attributes'],
            ['File Path', 'Relative path to the dataset file within the download directory'],
            ['Format', 'Data format (File Geodatabase, Shapefile, KML, GeoJSON, etc.)'],
            ['File Size', 'Size of the downloaded file'],
            ['Download Date', 'Date and time when the dataset was downloaded'],
            ['Source URL', 'Original URL where the data was obtained'],
            ['Coordinate System', 'Spatial reference system (EPSG code or WKID)'],
            ['Geometry Type', 'Type of geometry (Point, Polyline, Polygon, etc.)'],
            ['Feature Count', 'Number of features in the dataset (if available)']
        ]
        
        for row in field_descriptions:
            metadata_ws.append(row)
        
        # Style metadata sheet
        metadata_ws['A1'].fill = header_fill
        metadata_ws['A1'].font = header_font
        metadata_ws['B1'].fill = header_fill
        metadata_ws['B1'].font = header_font
        metadata_ws.column_dimensions['A'].width = 25
        metadata_ws.column_dimensions['B'].width = 70
        
        for row in metadata_ws.iter_rows(min_row=2, max_row=len(field_descriptions)+1):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Create README sheet
        readme_ws = wb.create_sheet("README", 0)  # Insert as first sheet
        readme_content = [
            ['BOEM Gulf of America Region (GOAR) Data Catalog'],
            [''],
            ['About This Catalog'],
            ['This Excel workbook catalogs all spatial and geographic data downloaded from the Bureau of Ocean Energy Management (BOEM) for the Gulf of America Region. The catalog provides detailed information about each dataset including descriptions, formats, file locations, and metadata.'],
            [''],
            ['Sheets in This Workbook'],
            ['• README - This sheet with overview information'],
            ['• Data Catalog - Complete listing of all downloaded datasets'],
            ['• Summary - Statistical summary of downloads by category and format'],
            ['• Field Descriptions - Detailed explanation of catalog fields'],
            [''],
            ['Data Sources'],
            ['• BOEM ArcGIS REST Services (https://gis.boem.gov/arcgis/rest/services)'],
            ['• BOEM Data Portal (https://www.data.boem.gov)'],
            ['• BOEM Gulf of America GIS Data (https://www.boem.gov/oil-gas-energy/mapping-and-data/goar-geographic-information-system-gis-data-and-maps)'],
            [''],
            ['Coordinate System'],
            ['Most BOEM Gulf data uses NAD 1927 (EPSG: 4267) as the standard coordinate system. Some newer datasets may use WGS 1984 (EPSG: 4326) or NAD 1983 (EPSG: 4269). Always verify the coordinate system before analysis.'],
            [''],
            ['Data Usage Notes'],
            ['• File Geodatabases (.gdb.zip) must be extracted before use in ArcGIS or QGIS'],
            ['• Shapefiles (.zip) contain multiple files - extract all components'],
            ['• KML/KMZ files can be opened directly in Google Earth or imported to GIS software'],
            ['• GeoJSON files are web-friendly and work with many modern mapping libraries'],
            [''],
            ['Data Currency'],
            ['Download Date:', datetime.now().strftime('%Y-%m-%d')],
            ['Note: BOEM updates data regularly. Check source URLs for the most current versions.'],
            [''],
            ['Contact Information'],
            ['For questions about BOEM data, visit: https://www.boem.gov'],
            ['For technical support: https://www.boem.gov/about-boem/contact-us'],
            [''],
            ['Disclaimer'],
            ['These data are provided "as is" from BOEM sources. Users are responsible for verifying data accuracy and fitness for their intended use. Official boundary coordinates are only those shown on Official Protraction Diagrams (OPDs) and Supplemental Official Block Diagrams (SOBDs).']
        ]
        
        for row in readme_content:
            readme_ws.append(row)
        
        # Style README
        readme_ws['A1'].font = Font(bold=True, size=16, color='0066CC')
        readme_ws.column_dimensions['A'].width = 120
        
        for row_num in [3, 6, 12, 17, 20, 24, 28, 31]:
            readme_ws[f'A{row_num}'].font = Font(bold=True, size=12)
        
        for row in readme_ws.iter_rows(min_row=1):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Save workbook
        wb.save(catalog_path)
        print(f"✓ Excel catalog created: {catalog_path}")
        print(f"  Total datasets cataloged: {len(self.data_catalog)}")
        print(f"  Categories: {len(category_counts)}")
        print(f"  Formats: {len(format_counts)}")
        
        return catalog_path
    
    def create_download_log(self):
        """Create a log file with download information."""
        log_file = self.output_dir / 'download_log.txt'
        
        with open(log_file, 'w') as f:
            f.write(f"BOEM GOAR Data Download Log\n")
            f.write(f"{'='*50}\n")
            f.write(f"Download Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Output Directory: {self.output_dir}\n\n")
            f.write(f"Data Sources:\n")
            f.write(f"  - ArcGIS REST Services\n")
            f.write(f"  - data.boem.gov Mapping Files\n")
            f.write(f"  - Cadastral and Boundary Data\n\n")
            f.write(f"Folder Structure:\n")
            
            for root, dirs, files in os.walk(self.output_dir):
                level = root.replace(str(self.output_dir), '').count(os.sep)
                indent = ' ' * 2 * level
                f.write(f"{indent}{os.path.basename(root)}/\n")
                sub_indent = ' ' * 2 * (level + 1)
                for file in files[:5]:  # Show first 5 files
                    f.write(f"{sub_indent}{file}\n")
                if len(files) > 5:
                    f.write(f"{sub_indent}... and {len(files)-5} more files\n")
        
        print(f"\n✓ Download log created: {log_file}")
    
    def run_full_download(self):
        """Execute full download of all GOAR data."""
        print(f"\n{'='*60}")
        print("BOEM Gulf of America Region Data Downloader")
        print(f"{'='*60}\n")
        print(f"Output directory: {self.output_dir}\n")
        
        try:
            # Download from different sources
            self.download_rest_services()
            self.download_cadastral_data()
            self.download_boundary_data()
            
            # Create Excel catalog
            catalog_path = self.create_excel_catalog()
            
            # Create summary log
            self.create_download_log()
            
            print(f"\n{'='*60}")
            print("Download Complete!")
            print(f"{'='*60}")
            print(f"\nAll data saved to: {self.output_dir}")
            print(f"Excel catalog: {catalog_path}")
            print("\nNote: Some URLs may have changed or require authentication.")
            print("Check the download log and Excel catalog for details.")
            
        except Exception as e:
            print(f"\nError during download: {str(e)}")
            raise


def main():
    """Main execution function."""
    print("\nBOEM GOAR Data Downloader with Excel Catalog")
    print("=" * 60)
    
    # Get output directory from user
    output_dir = input("\nEnter the path to your external hard drive or output folder: ").strip()
    
    if not output_dir:
        print("Error: No output directory specified.")
        return
    
    # Create subdirectory for BOEM data
    output_dir = os.path.join(output_dir, f"BOEM_GOAR_Data_{datetime.now().strftime('%Y%m%d')}")
    
    # Confirm with user
    print(f"\nData will be downloaded to: {output_dir}")
    print("An Excel catalog will be created with detailed descriptions of all datasets.")
    print("Includes 5 & 10 year lease term lines.")
    confirm = input("\nProceed with download? (yes/no): ").strip().lower()
    
    if confirm not in ['yes', 'y']:
        print("Download cancelled.")
        return
    
    # Check for openpyxl
    try:
        import openpyxl
    except ImportError:
        print("\nError: openpyxl is required for Excel catalog creation.")
        print("Install it with: pip install openpyxl")
        return
    
    # Initialize and run downloader
    downloader = BOEMDataDownloader(output_dir)
    downloader.run_full_download()


if __name__ == "__main__":
    main()