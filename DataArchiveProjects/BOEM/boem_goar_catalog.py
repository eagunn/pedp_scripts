"""
BOEM Gulf of America Region (GOAR) Data Catalog Generator
Creates a comprehensive Excel catalog of BOEM data without downloading files.
"""

import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path


class BOEMCatalogGenerator:
    def __init__(self, output_dir):
        """
        Initialize the catalog generator.
        
        Args:
            output_dir (str): Path to save the catalog
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.data_catalog = []
        
    def build_catalog_data(self):
        """Build the catalog with all BOEM GOAR datasets and descriptions."""
        
        # Cadastral and Leasing Data
        cadastral_datasets = [
            {
                'Dataset Name': 'Active Leases',
                'Category': 'Leasing',
                'Description': 'Polygons representing active oil and gas leases in the Gulf of America. Includes lease numbers, operators, effective dates, expiration dates, and lease status. Critical for identifying current leasing activity and operator information. Updated regularly as new leases are issued and existing leases expire.',
                'Format': 'File Geodatabase / Shapefile',
                'Source': 'BOEM Data Portal',
                'Source URL': 'https://www.data.boem.gov/Mapping/Files/',
                'Coordinate System': 'NAD 1927',
                'Geometry Type': 'Polygon',
                'Typical Attributes': 'Lease Number, Operator Name, Effective Date, Expiration Date, Status, Block Number, Area (acres)'
            },
            {
                'Dataset Name': 'Block Polygons',
                'Category': 'Cadastral',
                'Description': 'OCS block boundaries representing the cadastral grid system used for leasing. Each block is approximately 5,760 acres (3 miles x 3 miles). Contains block numbers, protraction areas, and official boundary coordinates. Foundation dataset for all Gulf cadastral operations.',
                'Format': 'File Geodatabase / Shapefile',
                'Source': 'BOEM Data Portal',
                'Source URL': 'https://www.data.boem.gov/Mapping/Files/',
                'Coordinate System': 'NAD 1927',
                'Geometry Type': 'Polygon',
                'Typical Attributes': 'Block Number, Protraction Name, Protraction Number, Area (acres), State Waters/Federal Waters'
            },
            {
                'Dataset Name': 'Protraction Diagrams',
                'Category': 'Cadastral',
                'Description': 'Protraction diagram boundaries that group OCS blocks into management units. Each protraction is typically 1 degree latitude by 2 degrees longitude. Used for planning and administrative purposes. Follows USGS topographic map series naming conventions.',
                'Format': 'File Geodatabase / Shapefile',
                'Source': 'BOEM Data Portal',
                'Source URL': 'https://www.data.boem.gov/Mapping/Files/',
                'Coordinate System': 'NAD 1927',
                'Geometry Type': 'Polygon',
                'Typical Attributes': 'Protraction Number, Protraction Name, Area, Number of Blocks'
            },
            {
                'Dataset Name': 'Official Protraction Diagrams (OPDs)',
                'Category': 'Cadastral',
                'Description': 'Official map documents showing block layouts, boundaries, and area measurements. Available as PDF images and GIS files. These are the legally binding documents for offshore coordinates. Critical reference for boundary disputes and lease descriptions.',
                'Format': 'PDF / GIS Files (.e00, Shapefile)',
                'Source': 'BOEM OPD Library',
                'Source URL': 'https://www.boem.gov/renewable-energy/mapping-and-data/official-protraction-diagrams-opds-and-leasing-maps-lms',
                'Coordinate System': 'NAD 1927',
                'Geometry Type': 'Mixed (Maps with cadastral features)',
                'Typical Attributes': 'Map series, blocks depicted, revision dates'
            },
            {
                'Dataset Name': 'Supplemental Official Block Diagrams (SOBDs)',
                'Category': 'Cadastral',
                'Description': 'Detailed diagrams for blocks intersected by offshore boundaries (SLA boundary, 8(g) zone, marine sanctuaries, etc.). Shows precise boundary intersections and split block areas. Essential for accurate area calculations and revenue sharing.',
                'Format': 'PDF',
                'Source': 'BOEM OPD Library',
                'Source URL': 'https://www.boem.gov/renewable-energy/mapping-and-data/official-protraction-diagrams-opds-and-leasing-maps-lms',
                'Coordinate System': 'NAD 1927',
                'Geometry Type': 'Diagram',
                'Typical Attributes': 'Block number, boundary intersections, area measurements'
            },
            {
                'Dataset Name': 'Lease Term Lines (5 & 10 Year)',
                'Category': 'Cadastral',
                'Description': 'Lines delineating lease term durations based on water depth. Leases seaward of the line have 10-year terms (deeper water), while leases landward have 5-year terms (shallow water). Critical for understanding lease duration and planning exploration/development timelines. Varies by lease sale.',
                'Format': 'Shapefile / GIS',
                'Source': 'BOEM Data Portal / Lease Sale Documents',
                'Source URL': 'https://www.data.boem.gov/Mapping/Files/',
                'Coordinate System': 'NAD 1927',
                'Geometry Type': 'Polyline',
                'Typical Attributes': 'Lease Sale Number, Term Duration, Effective Date'
            }
        ]
        
        # Infrastructure Data
        infrastructure_datasets = [
            {
                'Dataset Name': 'Oil and Gas Platforms',
                'Category': 'Infrastructure',
                'Description': 'Point locations of oil and gas platforms/structures in the Gulf. Includes structure names, complex IDs, installation dates, water depths, and operational status. Essential for infrastructure mapping, spatial planning, and conflict analysis. Updated regularly as platforms are installed or decommissioned.',
                'Format': 'File Geodatabase / Shapefile / KML',
                'Source': 'BOEM Data Portal',
                'Source URL': 'https://www.data.boem.gov/Mapping/Files/',
                'Coordinate System': 'NAD 1927',
                'Geometry Type': 'Point',
                'Typical Attributes': 'Structure Name, Complex ID, Install Date, Removal Date, Water Depth, Status, Operator, Block Number'
            },
            {
                'Dataset Name': 'Pipelines',
                'Category': 'Infrastructure',
                'Description': 'Pipeline routes and right-of-ways in the Gulf. Includes pipeline segments, operators, diameters, product types (oil, gas, multi-phase), and operational status. Critical for understanding offshore transportation networks, planning activities, and identifying potential conflicts with other uses.',
                'Format': 'File Geodatabase / Shapefile / KML',
                'Source': 'BOEM Data Portal',
                'Source URL': 'https://www.data.boem.gov/Mapping/Files/',
                'Coordinate System': 'NAD 1927',
                'Geometry Type': 'Polyline',
                'Typical Attributes': 'Pipeline Segment ID, Operator, Diameter, Product Type, Status, Approval Date, Block Numbers'
            },
            {
                'Dataset Name': 'Wells',
                'Category': 'Infrastructure',
                'Description': 'Well locations including exploratory, development, and abandoned wells. Contains API numbers, well names, spud dates, completion dates, total depths, and current status. Critical for resource assessment and development planning.',
                'Format': 'File Geodatabase / Shapefile',
                'Source': 'BOEM Data Portal / Well Data System',
                'Source URL': 'https://www.data.boem.gov/',
                'Coordinate System': 'NAD 1927',
                'Geometry Type': 'Point',
                'Typical Attributes': 'API Number, Well Name, Operator, Spud Date, Total Depth, Status, Production Type'
            }
        ]
        
        # Boundary Data
        boundary_datasets = [
            {
                'Dataset Name': 'State Seaward Boundaries (SLA)',
                'Category': 'Administrative Boundaries',
                'Description': 'Submerged Lands Act boundaries marking the division between state and federal waters. Typically 3 nautical miles from shore, but 9 nautical miles for Texas and Gulf Coast Florida (historical boundary). Critical for jurisdictional determinations, permitting, and revenue sharing.',
                'Format': 'File Geodatabase / Shapefile / KML',
                'Source': 'BOEM Data Portal',
                'Source URL': 'https://www.data.boem.gov/Mapping/Files/',
                'Coordinate System': 'NAD 1927',
                'Geometry Type': 'Polyline',
                'Typical Attributes': 'State, Boundary Type, Legal Description, Approved Date'
            },
            {
                'Dataset Name': '8(g) Zone Boundary',
                'Category': 'Administrative Boundaries',
                'Description': 'The 8(g) revenue sharing zone extending 3 miles seaward of state waters (typically 3-6 miles offshore, 9-12 miles for TX/FL). States receive 27% of revenues from leases in this zone per the Gulf of Mexico Energy Security Act (GOMESA). Important for fiscal analysis and state planning.',
                'Format': 'File Geodatabase / Shapefile',
                'Source': 'BOEM Data Portal',
                'Source URL': 'https://www.data.boem.gov/Mapping/Files/',
                'Coordinate System': 'NAD 1927',
                'Geometry Type': 'Polyline',
                'Typical Attributes': 'State, Zone Type, Legislation Reference'
            },
            {
                'Dataset Name': 'Planning Areas',
                'Category': 'Administrative Boundaries',
                'Description': 'BOEM planning area boundaries used for lease sale planning and resource management. Includes Western, Central, and Eastern Gulf planning areas. Used for regional analysis, Five-Year Program planning, and environmental assessments.',
                'Format': 'File Geodatabase / Shapefile / KML',
                'Source': 'BOEM Data Portal',
                'Source URL': 'https://www.data.boem.gov/Mapping/Files/',
                'Coordinate System': 'NAD 1927',
                'Geometry Type': 'Polygon',
                'Typical Attributes': 'Planning Area Name, Region, Area (sq km)'
            },
            {
                'Dataset Name': 'Fairways and Anchorage Areas',
                'Category': 'Maritime',
                'Description': 'Designated navigation fairways, shipping channels, and anchorage areas in the Gulf. Areas where oil and gas activities may be restricted or prohibited to ensure safe vessel passage. Maintained in coordination with USCG. Important for maritime planning, safety analysis, and activity permitting.',
                'Format': 'File Geodatabase / Shapefile',
                'Source': 'BOEM Data Portal',
                'Source URL': 'https://www.data.boem.gov/Mapping/Files/',
                'Coordinate System': 'NAD 1927',
                'Geometry Type': 'Polygon',
                'Typical Attributes': 'Fairway Name, Type, Restrictions, USCG Reference'
            },
            {
                'Dataset Name': 'Continental Shelf Boundary',
                'Category': 'Administrative Boundaries',
                'Description': 'Marks the limit of U.S. jurisdiction for offshore mineral development per international law. Defines the seaward extent of BOEM regulatory authority. Critical for understanding jurisdictional limits.',
                'Format': 'File Geodatabase / Shapefile',
                'Source': 'BOEM Data Portal',
                'Source URL': 'https://www.data.boem.gov/Mapping/Files/',
                'Coordinate System': 'NAD 1927',
                'Geometry Type': 'Polyline',
                'Typical Attributes': 'Boundary Type, Legal Basis'
            }
        ]
        
        # Environmental and Resource Data
        environmental_datasets = [
            {
                'Dataset Name': 'Topographic Features Stipulation Areas',
                'Category': 'Environmental Protection',
                'Description': 'Areas with special protective measures for topographic features (banks, reefs, etc.) that support diverse biological communities. Lease stipulations restrict certain activities to protect these sensitive features. Updated regularly based on new surveys and biological data.',
                'Format': 'PDF Maps / Shapefile',
                'Source': 'BOEM Gulf Region',
                'Source URL': 'https://www.boem.gov/newsroom/topographic-features-stipulation-map-package',
                'Coordinate System': 'NAD 1927',
                'Geometry Type': 'Polygon',
                'Typical Attributes': 'Feature Name, Stipulation Type, Effective Date, Restrictions'
            },
            {
                'Dataset Name': 'Coastal Zone Management (CZM) Blocks',
                'Category': 'Administrative',
                'Description': 'OCS blocks subject to state Coastal Zone Management Act review. Activities in these blocks require state consistency review. Interactive map application available for querying CZM blocks and viewing consistency requirements.',
                'Format': 'Interactive Map / Shapefile',
                'Source': 'BOEM ArcGIS Online',
                'Source URL': 'https://bobson.maps.arcgis.com/apps/webappviewer/',
                'Coordinate System': 'NAD 1927',
                'Geometry Type': 'Polygon',
                'Typical Attributes': 'Block Number, State CZM Authority, Review Requirements'
            },
            {
                'Dataset Name': 'Bathymetry and Water Depth',
                'Category': 'Environmental',
                'Description': 'Water depth contours and bathymetric data for the Gulf. Compiled from various sources including NOAA, industry surveys, and BOEM studies. Essential for resource assessment, engineering design, and environmental analysis.',
                'Format': 'Raster / Contour Shapefile',
                'Source': 'Multiple (NOAA, BOEM)',
                'Source URL': 'https://www.ncei.noaa.gov/maps/bathymetry/',
                'Coordinate System': 'NAD 1983 / WGS 1984',
                'Geometry Type': 'Raster / Polyline',
                'Typical Attributes': 'Depth (meters), Source, Survey Date, Resolution'
            }
        ]
        
        # REST Service Layers
        rest_service_datasets = [
            {
                'Dataset Name': 'BOEM_BSEE MMC Layers',
                'Category': 'REST Service',
                'Description': 'Comprehensive REST service providing access to multiple BOEM and BSEE (Bureau of Safety and Environmental Enforcement) data layers. Includes leases, blocks, infrastructure, boundaries, and more. Primary web service for accessing current BOEM spatial data programmatically.',
                'Format': 'ArcGIS REST Service',
                'Source': 'BOEM GIS Services',
                'Source URL': 'https://gis.boem.gov/arcgis/rest/services/BOEM_BSEE/MMC_Layers/MapServer',
                'Coordinate System': 'Various (WGS 1984, NAD 1927)',
                'Geometry Type': 'Multiple',
                'Typical Attributes': 'Varies by layer - see service metadata'
            },
            {
                'Dataset Name': 'Gulf of Mexico Layers',
                'Category': 'REST Service',
                'Description': 'Regional REST service specifically for Gulf of America data. Optimized for Gulf-specific queries and applications. May include additional regional layers not in national services.',
                'Format': 'ArcGIS REST Service',
                'Source': 'BOEM GIS Services',
                'Source URL': 'https://gis.boem.gov/arcgis/rest/services/BOEM_BSEE/GOM_Layers/MapServer',
                'Coordinate System': 'NAD 1927',
                'Geometry Type': 'Multiple',
                'Typical Attributes': 'Varies by layer - see service metadata'
            }
        ]
        
        # Add all datasets to catalog
        all_datasets = (cadastral_datasets + infrastructure_datasets + 
                       boundary_datasets + environmental_datasets + 
                       rest_service_datasets)
        
        self.data_catalog = all_datasets
        
    def create_excel_catalog(self):
        """Create a comprehensive Excel data catalog."""
        print("\n=== Creating BOEM GOAR Data Catalog ===\n")
        
        catalog_path = self.output_dir / 'BOEM_GOAR_Data_Catalog.xlsx'
        
        # Build catalog data
        self.build_catalog_data()
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data Catalog"
        
        # Define headers
        headers = [
            'Dataset Name',
            'Category',
            'Description',
            'Format',
            'Source',
            'Source URL',
            'Coordinate System',
            'Geometry Type',
            'Typical Attributes'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Add data
        for row_num, entry in enumerate(self.data_catalog, 2):
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = entry.get(header, '')
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Adjust column widths
        column_widths = {
            'A': 30,  # Dataset Name
            'B': 22,  # Category
            'C': 70,  # Description
            'D': 25,  # Format
            'E': 20,  # Source
            'F': 45,  # Source URL
            'G': 18,  # Coordinate System
            'H': 15,  # Geometry Type
            'I': 50   # Typical Attributes
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Add auto-filter
        ws.auto_filter.ref = ws.dimensions
        
        # Create summary sheet
        summary_ws = wb.create_sheet("Summary", 0)
        summary_ws.append(['BOEM Gulf of America Region Data Catalog'])
        summary_ws.append([''])
        summary_ws.append(['Catalog Information'])
        summary_ws.append(['Created Date:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        summary_ws.append(['Total Datasets:', len(self.data_catalog)])
        summary_ws.append([''])
        summary_ws.append(['Datasets by Category'])
        
        # Count by category
        category_counts = {}
        for entry in self.data_catalog:
            cat = entry.get('Category', 'Unknown')
            category_counts[cat] = category_counts.get(cat, 0) + 1
        
        for category, count in sorted(category_counts.items()):
            summary_ws.append([category, count])
        
        # Style summary sheet
        summary_ws['A1'].font = Font(bold=True, size=14, color='0066CC')
        summary_ws['A3'].font = Font(bold=True, size=12)
        summary_ws['A7'].font = Font(bold=True, size=12)
        summary_ws.column_dimensions['A'].width = 35
        summary_ws.column_dimensions['B'].width = 15
        
        # Create categories overview sheet
        categories_ws = wb.create_sheet("Categories Overview")
        categories_ws.append(['Category', 'Description', 'Key Datasets'])
        
        category_info = [
            ['Cadastral', 'Legal framework for offshore leasing including blocks, protraction diagrams, and official maps', 'Blocks, Protraction Diagrams, OPDs, SOBDs'],
            ['Leasing', 'Active and historical lease information showing operator activity and lease status', 'Active Leases, Lease History'],
            ['Infrastructure', 'Physical structures including platforms, pipelines, and wells', 'Platforms, Pipelines, Wells'],
            ['Administrative Boundaries', 'Legal and jurisdictional boundaries for management and regulation', 'State Waters, 8(g) Zone, Planning Areas'],
            ['Maritime', 'Navigation and shipping related features', 'Fairways, Anchorage Areas'],
            ['Environmental Protection', 'Areas with special protective measures', 'Topographic Features, Protected Areas'],
            ['REST Service', 'Web services providing programmatic access to data', 'MMC Layers, GOM Layers']
        ]
        
        for row in category_info:
            categories_ws.append(row)
        
        # Style categories sheet
        categories_ws['A1'].fill = header_fill
        categories_ws['A1'].font = header_font
        categories_ws['B1'].fill = header_fill
        categories_ws['B1'].font = header_font
        categories_ws['C1'].fill = header_fill
        categories_ws['C1'].font = header_font
        categories_ws.column_dimensions['A'].width = 25
        categories_ws.column_dimensions['B'].width = 60
        categories_ws.column_dimensions['C'].width = 40
        
        for row in categories_ws.iter_rows(min_row=2, max_row=len(category_info)+1):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Create README sheet
        readme_ws = wb.create_sheet("README", 0)
        readme_content = [
            ['BOEM Gulf of America Region (GOAR) Data Catalog'],
            [''],
            ['About This Catalog'],
            ['This Excel workbook provides a comprehensive catalog of all spatial and geographic data available from the Bureau of Ocean Energy Management (BOEM) for the Gulf of America Region. The catalog includes detailed descriptions of each dataset, their formats, sources, and typical attributes.'],
            [''],
            ['Purpose'],
            ['This catalog serves as a reference guide to help users:'],
            ['• Understand what BOEM data is available for the Gulf of America'],
            ['• Learn about the content and attributes of each dataset'],
            ['• Identify the appropriate data sources for their needs'],
            ['• Plan data acquisition and integration strategies'],
            [''],
            ['Sheets in This Workbook'],
            ['• README - Overview and guidance (this sheet)'],
            ['• Summary - Statistical overview of datasets'],
            ['• Categories Overview - Explanation of data categories'],
            ['• Data Catalog - Complete detailed listing of all datasets'],
            [''],
            ['How to Use This Catalog'],
            ['1. Start with the Summary sheet to understand the scope of available data'],
            ['2. Review the Categories Overview to understand data organization'],
            ['3. Use the Data Catalog sheet to find specific datasets'],
            ['4. Use filters and search to find datasets relevant to your work'],
            ['5. Visit the Source URLs to access and download actual data files'],
            [''],
            ['Data Sources'],
            ['• BOEM Data Portal: https://www.data.boem.gov'],
            ['• BOEM ArcGIS REST Services: https://gis.boem.gov/arcgis/rest/services'],
            ['• BOEM Gulf GIS Page: https://www.boem.gov/oil-gas-energy/mapping-and-data/goar-geographic-information-system-gis-data-and-maps'],
            ['• Marine Cadastre: https://marinecadastre.gov'],
            [''],
            ['Coordinate Systems'],
            ['Most BOEM Gulf data uses NAD 1927 (EPSG: 4267) as the standard coordinate system due to historical continuity with the official cadastral framework. Some newer datasets may use NAD 1983 (EPSG: 4269) or WGS 1984 (EPSG: 4326). Always verify the coordinate system before integrating data.'],
            [''],
            ['Important Notes'],
            ['• This is a CATALOG ONLY - it does not contain the actual data files'],
            ['• Data must be downloaded separately from the listed source URLs'],
            ['• BOEM updates data regularly - check source URLs for the most current versions'],
            ['• Some data requires specific GIS software to use (ArcGIS, QGIS, etc.)'],
            ['• Official boundary coordinates are only those shown on Official Protraction Diagrams (OPDs) and Supplemental Official Block Diagrams (SOBDs)'],
            [''],
            ['Data Currency'],
            ['Catalog Created:', datetime.now().strftime('%Y-%m-%d')],
            ['Note: Dataset descriptions reflect data structure and content as of the catalog creation date. Always check BOEM sources for updates.'],
            [''],
            ['Contact Information'],
            ['BOEM Website: https://www.boem.gov'],
            ['BOEM Contact: https://www.boem.gov/about-boem/contact-us'],
            ['Gulf Regional Office: 1201 Elmwood Park Blvd, New Orleans, LA 70123'],
            ['Phone: 1-800-200-GULF'],
            [''],
            ['Disclaimer'],
            ['This catalog is provided for informational purposes. Users should verify data accuracy, currency, and fitness for their intended use. Official records are maintained by BOEM. For legal or official purposes, consult BOEM directly.']
        ]
        
        for row in readme_content:
            readme_ws.append(row)
        
        # Style README
        readme_ws['A1'].font = Font(bold=True, size=16, color='0066CC')
        readme_ws.column_dimensions['A'].width = 120
        
        # Bold section headers
        bold_rows = [3, 6, 13, 19, 26, 31, 37, 42, 47]
        for row_num in bold_rows:
            readme_ws[f'A{row_num}'].font = Font(bold=True, size=12)
        
        for row in readme_ws.iter_rows(min_row=1):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Save workbook
        wb.save(catalog_path)
        print(f"✓ Excel catalog created: {catalog_path}")
        print(f"  Total datasets cataloged: {len(self.data_catalog)}")
        print(f"  Categories: {len(category_counts)}")
        
        return catalog_path


def main():
    """Main execution function."""
    print("\nBOEM GOAR Data Catalog Generator")
    print("=" * 60)
    print("\nThis script creates a comprehensive Excel catalog of BOEM Gulf")
    print("of America Region datasets WITHOUT downloading the actual data files.")
    print("=" * 60)
    
    # Get output directory from user
    output_dir = input("\nEnter the path where you want to save the catalog: ").strip()
    
    if not output_dir:
        output_dir = "."
        print(f"Using current directory: {os.path.abspath(output_dir)}")
    
    # Confirm with user
    print(f"\nCatalog will be saved to: {os.path.abspath(output_dir)}")
    confirm = input("\nProceed? (yes/no): ").strip().lower()
    
    if confirm not in ['yes', 'y']:
        print("Cancelled.")
        return
    
    # Check for openpyxl
    try:
        import openpyxl
    except ImportError:
        print("\nError: openpyxl is required for Excel catalog creation.")
        print("Install it with: pip install openpyxl")
        return
    
    # Generate catalog
    generator = BOEMCatalogGenerator(output_dir)
    catalog_path = generator.create_excel_catalog()
    
    print("\n" + "=" * 60)
    print("Catalog Generation Complete!")
    print("=" * 60)
    print(f"\nCatalog saved to: {catalog_path}")
    print("\nThe catalog contains:")
    print("  • README with usage instructions")
    print("  • Summary statistics")
    print("  • Category descriptions")
    print("  • Detailed dataset catalog with descriptions")
    print("\nUse this catalog to identify datasets, then download them")
    print("from the source URLs listed in the catalog.")


if __name__ == "__main__":
    main()
